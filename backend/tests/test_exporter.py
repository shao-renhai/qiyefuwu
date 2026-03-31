import sys
import os
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))


SAMPLE_DATA = {
    "client": {"name": "张三", "created_at": "2026-03-31"},
    "credit": {
        "total_debt": 1285000,
        "total_balance": 1285000,
        "institution_details": [{"type": "住房贷款", "balance": 850000}],
        "credit_card_total_limit": 500000,
        "credit_card_used": 235000,
        "credit_card_usage_rate": 47.0,
        "active_loans": [{"type": "住房贷款", "balance": 850000}],
        "overdue_records": [{"type": "历史逾期", "count": 1}],
        "query_records": {
            "recent_1m": {"loan_approval": 1, "corporate_review": 0},
            "recent_3m": {"loan_approval": 3, "corporate_review": 1},
            "recent_6m": {"loan_approval": 5, "corporate_review": 1},
            "recent_1y": {"loan_approval": 7, "corporate_review": 2},
        },
    },
    "bank": {
        "total_income": 600000,
        "total_expense": 300000,
        "monthly_avg_income": 100000,
        "monthly_avg_expense": 50000,
        "monthly_avg_net": 50000,
        "deduped_total_income": 500000,
        "deduped_total_expense": 250000,
        "deduped_monthly_avg_income": 83333,
        "deduped_monthly_avg_expense": 41667,
        "top_income_sources": [
            {"counterparty": "某公司", "amount": 300000, "ratio": 50.0}
        ],
        "top_expense_categories": [
            {"counterparty": "供应商A", "amount": 100000, "ratio": 33.3}
        ],
        "monthly_ending_balances": [
            {"month": "2026-01", "balance": 200000},
            {"month": "2026-02", "balance": 250000},
        ],
        "min_balance": 50000,
        "avg_balance": 150000,
        "monthly_avg_tx_count": 30,
        "daily_avg_tx_count": 1.5,
        "monthly_summary": [
            {
                "month": "2026-01",
                "income": 300000,
                "expense": 150000,
                "net": 150000,
                "tx_count": 30,
            },
            {
                "month": "2026-02",
                "income": 300000,
                "expense": 150000,
                "net": 150000,
                "tx_count": 30,
            },
        ],
        "anomalies": [
            {
                "date": "2026-01-15",
                "counterparty": "某人",
                "amount": 500000,
                "direction": "收入",
                "type": "large_amount",
                "description": "单笔金额500,000元，超过月均收入2倍",
            },
        ],
    },
}


def test_export_excel(tmp_path):
    from services.exporter import export_excel

    filepath = str(tmp_path / "test_report.xlsx")
    export_excel(SAMPLE_DATA, filepath)
    assert os.path.exists(filepath)
    assert os.path.getsize(filepath) > 0
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    assert "客户概览" in wb.sheetnames
    assert "征信详情" in wb.sheetnames
    assert "流水汇总" in wb.sheetnames
    assert "异常交易" in wb.sheetnames


def test_export_pdf(tmp_path):
    from services.exporter import export_pdf

    filepath = str(tmp_path / "test_report.pdf")
    export_pdf(SAMPLE_DATA, filepath)
    assert os.path.exists(filepath)
    assert os.path.getsize(filepath) > 0
