"""
Export service: generates Excel and PDF reports from analysis data.
"""

import os
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers


# --- Styling constants ---
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)


def _fmt_money(amount: float) -> str:
    """Format amount as '128.50万' if >=10000, else 'xxx.xx元'."""
    if amount is None:
        return "0.00元"
    amount = float(amount)
    if abs(amount) >= 10000:
        return f"{amount / 10000:.2f}万"
    return f"{amount:.2f}元"


def _write_header_row(ws, row: int, headers: list):
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _write_data_row(ws, row: int, values: list):
    """Write a data row with borders."""
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = THIN_BORDER


def _write_label_value(ws, row: int, label: str, value: str, col_start: int = 1):
    """Write a label-value pair."""
    label_cell = ws.cell(row=row, column=col_start, value=label)
    label_cell.font = Font(bold=True)
    label_cell.border = THIN_BORDER
    value_cell = ws.cell(row=row, column=col_start + 1, value=value)
    value_cell.border = THIN_BORDER


def export_excel(data: Dict[str, Any], filepath: str):
    """Export analysis data to an Excel file with 4 sheets."""
    wb = Workbook()
    client = data.get("client", {})
    credit = data.get("credit", {})
    bank = data.get("bank", {})

    # ========== Sheet 1: 客户概览 ==========
    ws1 = wb.active
    ws1.title = "客户概览"
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 30

    ws1.cell(row=1, column=1, value="客户融资分析报告").font = TITLE_FONT
    ws1.merge_cells("A1:B1")

    _write_label_value(ws1, 3, "客户姓名", client.get("name", ""))
    _write_label_value(ws1, 4, "报告日期", client.get("created_at", ""))

    ws1.cell(row=6, column=1, value="关键指标").font = SECTION_FONT
    _write_label_value(ws1, 7, "负债总额", _fmt_money(credit.get("total_balance", 0)))
    _write_label_value(ws1, 8, "信用卡使用率", f"{credit.get('credit_card_usage_rate', 0):.1f}%")
    _write_label_value(ws1, 9, "月均收入(原始)", _fmt_money(bank.get("monthly_avg_income", 0)))
    _write_label_value(ws1, 10, "月均支出(原始)", _fmt_money(bank.get("monthly_avg_expense", 0)))
    _write_label_value(ws1, 11, "月均净收入", _fmt_money(bank.get("monthly_avg_net", 0)))
    _write_label_value(ws1, 12, "月均收入(去重)", _fmt_money(bank.get("deduped_monthly_avg_income", 0)))
    _write_label_value(ws1, 13, "最低余额", _fmt_money(bank.get("min_balance", 0)))
    _write_label_value(ws1, 14, "平均余额", _fmt_money(bank.get("avg_balance", 0)))

    # ========== Sheet 2: 征信详情 ==========
    ws2 = wb.create_sheet("征信详情")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 20

    ws2.cell(row=1, column=1, value="负债概况").font = SECTION_FONT
    _write_label_value(ws2, 2, "负债总额", _fmt_money(credit.get("total_debt", 0)))
    _write_label_value(ws2, 3, "未结余额", _fmt_money(credit.get("total_balance", 0)))

    # Institution details
    row = 5
    ws2.cell(row=row, column=1, value="机构明细").font = SECTION_FONT
    row += 1
    _write_header_row(ws2, row, ["贷款类型", "余额"])
    row += 1
    for item in credit.get("institution_details", []):
        _write_data_row(ws2, row, [item.get("type", ""), _fmt_money(item.get("balance", 0))])
        row += 1

    # Credit card summary
    row += 1
    ws2.cell(row=row, column=1, value="信用卡概况").font = SECTION_FONT
    row += 1
    _write_label_value(ws2, row, "授信总额", _fmt_money(credit.get("credit_card_total_limit", 0)))
    row += 1
    _write_label_value(ws2, row, "已用额度", _fmt_money(credit.get("credit_card_used", 0)))
    row += 1
    _write_label_value(ws2, row, "使用率", f"{credit.get('credit_card_usage_rate', 0):.1f}%")
    row += 1

    # Active loans
    row += 1
    ws2.cell(row=row, column=1, value="在贷明细").font = SECTION_FONT
    row += 1
    _write_header_row(ws2, row, ["贷款类型", "余额"])
    row += 1
    for loan in credit.get("active_loans", []):
        _write_data_row(ws2, row, [loan.get("type", ""), _fmt_money(loan.get("balance", 0))])
        row += 1

    # Overdue records
    row += 1
    ws2.cell(row=row, column=1, value="逾期记录").font = SECTION_FONT
    row += 1
    _write_header_row(ws2, row, ["类型", "次数"])
    row += 1
    for rec in credit.get("overdue_records", []):
        _write_data_row(ws2, row, [rec.get("type", ""), rec.get("count", 0)])
        row += 1

    # Query records by period
    row += 1
    ws2.cell(row=row, column=1, value="查询记录").font = SECTION_FONT
    row += 1
    _write_header_row(ws2, row, ["时间段", "贷款审批", "法人审查"])
    row += 1
    query_records = credit.get("query_records", {})
    period_labels = {
        "recent_1m": "近1个月",
        "recent_3m": "近3个月",
        "recent_6m": "近6个月",
        "recent_1y": "近1年",
    }
    for key, label in period_labels.items():
        period = query_records.get(key, {})
        _write_data_row(ws2, row, [label, period.get("loan_approval", 0), period.get("corporate_review", 0)])
        row += 1

    # ========== Sheet 3: 流水汇总 ==========
    ws3 = wb.create_sheet("流水汇总")
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 18

    ws3.cell(row=1, column=1, value="收支对比（原始 vs 去重）").font = SECTION_FONT
    _write_header_row(ws3, 2, ["指标", "原始", "去重后"])
    _write_data_row(ws3, 3, ["总收入", _fmt_money(bank.get("total_income", 0)), _fmt_money(bank.get("deduped_total_income", 0))])
    _write_data_row(ws3, 4, ["总支出", _fmt_money(bank.get("total_expense", 0)), _fmt_money(bank.get("deduped_total_expense", 0))])
    _write_data_row(ws3, 5, ["月均收入", _fmt_money(bank.get("monthly_avg_income", 0)), _fmt_money(bank.get("deduped_monthly_avg_income", 0))])
    _write_data_row(ws3, 6, ["月均支出", _fmt_money(bank.get("monthly_avg_expense", 0)), _fmt_money(bank.get("deduped_monthly_avg_expense", 0))])

    row = 8
    ws3.cell(row=row, column=1, value="余额与频次").font = SECTION_FONT
    row += 1
    _write_label_value(ws3, row, "最低余额", _fmt_money(bank.get("min_balance", 0)))
    row += 1
    _write_label_value(ws3, row, "平均余额", _fmt_money(bank.get("avg_balance", 0)))
    row += 1
    _write_label_value(ws3, row, "月均交易笔数", str(bank.get("monthly_avg_tx_count", 0)))
    row += 1
    _write_label_value(ws3, row, "日均交易笔数", f"{bank.get('daily_avg_tx_count', 0):.1f}")

    # Monthly breakdown
    row += 2
    ws3.cell(row=row, column=1, value="月度明细").font = SECTION_FONT
    row += 1
    _write_header_row(ws3, row, ["月份", "收入", "支出", "净收入", "交易笔数"])
    row += 1
    for item in bank.get("monthly_summary", []):
        _write_data_row(ws3, row, [
            item.get("month", ""),
            _fmt_money(item.get("income", 0)),
            _fmt_money(item.get("expense", 0)),
            _fmt_money(item.get("net", 0)),
            item.get("tx_count", 0),
        ])
        row += 1

    # Top income sources
    row += 1
    ws3.cell(row=row, column=1, value="主要收入来源").font = SECTION_FONT
    row += 1
    _write_header_row(ws3, row, ["交易对手", "金额", "占比"])
    row += 1
    for src in bank.get("top_income_sources", []):
        _write_data_row(ws3, row, [
            src.get("counterparty", ""),
            _fmt_money(src.get("amount", src.get("total", 0))),
            f"{src.get('ratio', 0):.1f}%",
        ])
        row += 1

    # Top expense categories
    row += 1
    ws3.cell(row=row, column=1, value="主要支出去向").font = SECTION_FONT
    row += 1
    _write_header_row(ws3, row, ["交易对手", "金额", "占比"])
    row += 1
    for cat in bank.get("top_expense_categories", []):
        _write_data_row(ws3, row, [
            cat.get("counterparty", ""),
            _fmt_money(cat.get("amount", cat.get("total", 0))),
            f"{cat.get('ratio', 0):.1f}%",
        ])
        row += 1

    # ========== Sheet 4: 异常交易 ==========
    ws4 = wb.create_sheet("异常交易")
    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 12
    ws4.column_dimensions["E"].width = 15
    ws4.column_dimensions["F"].width = 40

    ws4.cell(row=1, column=1, value="异常交易列表").font = SECTION_FONT
    _write_header_row(ws4, 2, ["日期", "交易对手", "金额", "方向", "类型", "说明"])
    row = 3
    for anomaly in bank.get("anomalies", []):
        _write_data_row(ws4, row, [
            anomaly.get("date", ""),
            anomaly.get("counterparty", ""),
            _fmt_money(anomaly.get("amount", 0)),
            anomaly.get("direction", ""),
            anomaly.get("type", ""),
            anomaly.get("description", anomaly.get("detail", "")),
        ])
        row += 1

    wb.save(filepath)


def _register_chinese_font():
    """Register a Chinese font for PDF generation. Returns font name."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    # Try system fonts in order
    font_paths = [
        ("/System/Library/Fonts/PingFang.ttc", "PingFang"),
        ("/System/Library/Fonts/STHeiti Light.ttc", "STHeiti"),
        ("/System/Library/Fonts/Hiragino Sans GB.ttc", "HiraginoSansGB"),
    ]

    for path, name in font_paths:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                return name
            except Exception:
                continue

    # Fallback: try CID font for Chinese
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        return "STSong-Light"
    except Exception:
        pass

    return "Helvetica"


def export_pdf(data: Dict[str, Any], filepath: str):
    """Export analysis data to a PDF report."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    font_name = _register_chinese_font()

    doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()

    # Create Chinese-compatible styles
    title_style = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=22,
        spaceAfter=10 * mm,
    )
    heading_style = ParagraphStyle(
        "ChineseHeading",
        parent=styles["Heading2"],
        fontName=font_name,
        fontSize=14,
        spaceBefore=8 * mm,
        spaceAfter=4 * mm,
    )
    normal_style = ParagraphStyle(
        "ChineseNormal",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
    )

    elements = []
    client = data.get("client", {})
    credit = data.get("credit", {})
    bank = data.get("bank", {})

    # Table style
    table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])

    # ========== Cover page ==========
    elements.append(Spacer(1, 60 * mm))
    elements.append(Paragraph("客户融资分析报告", title_style))
    elements.append(Spacer(1, 10 * mm))
    elements.append(Paragraph(f"客户: {client.get('name', '')}", normal_style))
    elements.append(Paragraph(f"日期: {client.get('created_at', '')}", normal_style))
    elements.append(PageBreak())

    # ========== 征信概况 ==========
    elements.append(Paragraph("征信概况", heading_style))
    credit_table_data = [
        ["指标", "数值"],
        ["负债总额", _fmt_money(credit.get("total_debt", 0))],
        ["未结余额", _fmt_money(credit.get("total_balance", 0))],
        ["信用卡授信总额", _fmt_money(credit.get("credit_card_total_limit", 0))],
        ["信用卡已用额度", _fmt_money(credit.get("credit_card_used", 0))],
        ["信用卡使用率", f"{credit.get('credit_card_usage_rate', 0):.1f}%"],
    ]
    # Add active loans
    for loan in credit.get("active_loans", []):
        credit_table_data.append([
            f"在贷: {loan.get('type', '')}",
            _fmt_money(loan.get("balance", 0)),
        ])
    # Add overdue records
    for rec in credit.get("overdue_records", []):
        credit_table_data.append([
            f"逾期: {rec.get('type', '')}",
            str(rec.get("count", 0)),
        ])

    t = Table(credit_table_data, colWidths=[80 * mm, 80 * mm])
    t.setStyle(table_style)
    elements.append(t)

    # ========== Query records ==========
    elements.append(Paragraph("查询记录", heading_style))
    query_records = credit.get("query_records", {})
    query_table_data = [["时间段", "贷款审批", "法人审查"]]
    period_labels = {
        "recent_1m": "近1个月",
        "recent_3m": "近3个月",
        "recent_6m": "近6个月",
        "recent_1y": "近1年",
    }
    for key, label in period_labels.items():
        period = query_records.get(key, {})
        query_table_data.append([
            label,
            str(period.get("loan_approval", 0)),
            str(period.get("corporate_review", 0)),
        ])

    t = Table(query_table_data, colWidths=[53 * mm, 53 * mm, 54 * mm])
    t.setStyle(table_style)
    elements.append(t)

    # ========== 银行流水分析 ==========
    elements.append(Paragraph("银行流水分析", heading_style))
    bank_table_data = [
        ["指标", "原始", "去重后"],
        ["总收入", _fmt_money(bank.get("total_income", 0)), _fmt_money(bank.get("deduped_total_income", 0))],
        ["总支出", _fmt_money(bank.get("total_expense", 0)), _fmt_money(bank.get("deduped_total_expense", 0))],
        ["月均收入", _fmt_money(bank.get("monthly_avg_income", 0)), _fmt_money(bank.get("deduped_monthly_avg_income", 0))],
        ["月均支出", _fmt_money(bank.get("monthly_avg_expense", 0)), _fmt_money(bank.get("deduped_monthly_avg_expense", 0))],
    ]

    t = Table(bank_table_data, colWidths=[53 * mm, 53 * mm, 54 * mm])
    t.setStyle(table_style)
    elements.append(t)

    # Additional bank metrics
    elements.append(Spacer(1, 4 * mm))
    bank_extra = [
        ["指标", "数值"],
        ["最低余额", _fmt_money(bank.get("min_balance", 0))],
        ["平均余额", _fmt_money(bank.get("avg_balance", 0))],
        ["月均交易笔数", str(bank.get("monthly_avg_tx_count", 0))],
        ["日均交易笔数", f"{bank.get('daily_avg_tx_count', 0):.1f}"],
    ]
    t = Table(bank_extra, colWidths=[80 * mm, 80 * mm])
    t.setStyle(table_style)
    elements.append(t)

    # ========== 异常交易 ==========
    anomalies = bank.get("anomalies", [])
    if anomalies:
        elements.append(Paragraph("异常交易", heading_style))
        anomaly_table_data = [["日期", "交易对手", "金额", "方向", "类型", "说明"]]
        for a in anomalies:
            anomaly_table_data.append([
                a.get("date", ""),
                a.get("counterparty", ""),
                _fmt_money(a.get("amount", 0)),
                a.get("direction", ""),
                a.get("type", ""),
                a.get("description", a.get("detail", "")),
            ])

        col_widths = [22 * mm, 25 * mm, 22 * mm, 15 * mm, 22 * mm, 54 * mm]
        t = Table(anomaly_table_data, colWidths=col_widths)
        t.setStyle(table_style)
        elements.append(t)

    doc.build(elements)
